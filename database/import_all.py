# database/import_all.py
import os, csv, json, hashlib, sys, re
from pathlib import Path

import psycopg2, psycopg2.extras

# ---------- Optional deps ----------
try:
    import phonenumbers
except Exception:
    phonenumbers = None

try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None

# ---------- ENV / CONFIG ----------
try:
    from dotenv import load_dotenv
    load_dotenv()
except Exception:
    pass

# CLI arg > .env > default
ROOT = Path(
    sys.argv[1] if len(sys.argv) > 1
    else os.getenv("ROOT", r"D:\Program Files (x86)\USA Database Business Leads-20251102T151819Z-1-001")
)
DB_URL = os.getenv("DATABASE_URL", "postgresql://dialer:dialer@localhost:5432/dialer")
BATCH_SIZE = int(os.getenv("BATCH_SIZE", "2000"))

# Heuristic column keys (case-insensitive)
NAME_KEYS    = {"business name","company","name","title"}
PHONE_KEYS   = {"phone","phone number","telephone","mobile","contact"}
EMAIL_KEYS   = {"email","e-mail","e mail"}
WEBSITE_KEYS = {"website","url","site"}
ADDR_KEYS    = {"address","street","street address","addr"}
CITY_KEYS    = {"city","town"}
STATE_KEYS   = {"state","region","province"}
ZIP_KEYS     = {"zip","zipcode","postal","postal code"}
CAT_KEYS     = {"category","type","industry"}  # reserved for future

# Optional file-name hint like ..._in_Austin_Texas.csv|xlsx
FILE_RE = re.compile(r"_in_([A-Za-z_]+)_([A-Za-z_]+)\.(csv|xlsx)$", re.IGNORECASE)

# ---------- Helpers ----------
def sha1_file(p: Path) -> str:
    h = hashlib.sha1()
    with p.open('rb') as f:
        for chunk in iter(lambda: f.read(1 << 16), b''):
            h.update(chunk)
    return h.hexdigest()

def slugify(name: str) -> str:
    s = re.sub(r"[^a-z0-9]+", "-", (name or "").lower()).strip("-")
    return s[:80] or "campaign"

def ensure_campaign(cur, name: str) -> int:
    """
    Reuse existing by exact name; otherwise create with a unique slug.
    If slug collision with a different name, append -2, -3, ...
    """
    base_slug = slugify(name)

    # by name
    cur.execute("SELECT id FROM campaigns WHERE name = %s", (name,))
    row = cur.fetchone()
    if row:
        return row[0]

    # by slug
    cur.execute("SELECT id, name FROM campaigns WHERE slug = %s", (base_slug,))
    r = cur.fetchone()
    if r:
        existing_id, existing_name = r
        if existing_name == name:
            return existing_id
        # disambiguate
        n = 2
        slug = f"{base_slug}-{n}"
        while True:
            cur.execute("SELECT 1 FROM campaigns WHERE slug = %s", (slug,))
            if not cur.fetchone():
                break
            n += 1
            slug = f"{base_slug}-{n}"
    else:
        slug = base_slug

    cur.execute(
        "INSERT INTO campaigns (name, slug) VALUES (%s,%s) RETURNING id",
        (name, slug),
    )
    return cur.fetchone()[0]

def tz_for(cur, city, state):
    if not city or not state:
        return None
    cur.execute("SELECT tz_name FROM cities WHERE city=%s AND state=%s", (city, state))
    r = cur.fetchone()
    return r[0] if r else None

def parse_city_state_from_file(name):
    m = FILE_RE.search(name)
    if not m:
        return (None, None)
    city = m.group(1).replace("_"," ").strip()
    state = m.group(2).replace("_"," ").strip()
    return (city, state)

def norm_phone(us_str):
    if not us_str:
        return None
    s = re.sub(r"[^0-9+]", "", str(us_str))
    if not s:
        return None
    if not phonenumbers:
        # fallback: return cleaned digits (not validated)
        return s
    try:
        num = phonenumbers.parse(s, "US")
        if phonenumbers.is_valid_number(num):
            return phonenumbers.format_number(num, phonenumbers.PhoneNumberFormat.E164)
    except Exception:
        return None
    return None

def keypick(row, keys):
    # row keys are case-insensitive
    for k in row.keys():
        if k is None:
            continue
        kk = str(k).strip().lower()
        if kk in keys:
            v = row[k]
            if v is None:
                continue
            v = str(v).strip()
            if v:
                return v
    return None

def upsert_lead_allow_nulls(cur, row_data):
    """
    Always insert even if phone/email missing.
    Dedup using actual UNIQUE CONSTRAINTS (not partial indexes).
    """
    # 1) Try by phone if present
    if row_data.get("phone_e164"):
        cur.execute("""
          INSERT INTO leads
            (company, contact, email, phone_raw, phone_e164, website, address,
             city, state, postal_code, tz_name, source_file, source_row, last_seen)
          VALUES (%(company)s,%(contact)s,%(email)s,%(phone_raw)s,%(phone_e164)s,%(website)s,%(address)s,
                  %(city)s,%(state)s,%(postal_code)s,%(tz_name)s,%(source_file)s,%(source_row)s, now())
          ON CONFLICT ON CONSTRAINT leads_phone_unique_constraint
          DO UPDATE SET last_seen = now(), source_file = EXCLUDED.source_file
          RETURNING id
        """, row_data)
        if cur.rowcount:
            return cur.fetchone()[0]

    # 2) Try by email if present
    if row_data.get("email"):
        cur.execute("""
          INSERT INTO leads
            (company, contact, email, phone_raw, phone_e164, website, address,
             city, state, postal_code, tz_name, source_file, source_row, last_seen)
          VALUES (%(company)s,%(contact)s,%(email)s,%(phone_raw)s,%(phone_e164)s,%(website)s,%(address)s,
                  %(city)s,%(state)s,%(postal_code)s,%(tz_name)s,%(source_file)s,%(source_row)s, now())
          ON CONFLICT ON CONSTRAINT leads_email_unique_constraint
          DO UPDATE SET last_seen = now(), source_file = EXCLUDED.source_file
          RETURNING id
        """, row_data)
        if cur.rowcount:
            return cur.fetchone()[0]

    # 3) Otherwise unconditional insert
    cur.execute("""
      INSERT INTO leads
        (company, contact, email, phone_raw, phone_e164, website, address,
         city, state, postal_code, tz_name, source_file, source_row, last_seen)
      VALUES (%(company)s,%(contact)s,%(email)s,%(phone_raw)s,%(phone_e164)s,%(website)s,%(address)s,
              %(city)s,%(state)s,%(postal_code)s,%(tz_name)s,%(source_file)s,%(source_row)s, now())
      RETURNING id
    """, row_data)
    return cur.fetchone()[0]



# ---------- Readers ----------
def dict_rows_from_csv(path: Path):
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        for r in reader:
            yield r

def dict_rows_from_xlsx(path: Path):
    if not load_workbook:
        raise RuntimeError("openpyxl is not installed. Run: pip install openpyxl")
    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    if not wb.worksheets:
        return
    ws = wb.worksheets[0]  # first sheet
    rows = ws.iter_rows(values_only=True)

    try:
        header = next(rows)
    except StopIteration:
        return

    header = [(str(h).strip() if h is not None else "") for h in header]
    if not any(header):
        return

    for row in rows:
        # row may be shorter than header
        r = { header[i]: (row[i] if i < len(row) else None) for i in range(len(header)) }
        yield r

# ---------- Processing ----------
def process_file(cur, file_path: Path):
    # Campaign = file name (stem)
    campaign_name = file_path.stem.strip()
    campaign_id = ensure_campaign(cur, campaign_name)

    # Skip duplicates by file hash if already done
    file_sha = sha1_file(file_path)
    cur.execute("SELECT 1 FROM imports WHERE file_sha1=%s AND status='done' LIMIT 1", (file_sha,))
    if cur.fetchone():
        print(f"  - SKIP already imported: {file_path.name}")
        return

    # Create import record
    cur.execute("""
        INSERT INTO imports (campaign_id, list_path, file_name, file_sha1, status)
        VALUES (%s,%s,%s,%s,'running') RETURNING id
    """, (campaign_id, str(file_path.parent), file_path.name, file_sha))
    imp_id = cur.fetchone()[0]

    # Parse optional city/state hint from the file name
    file_city, file_state = parse_city_state_from_file(file_path.name)

    # Choose reader
    ext = file_path.suffix.lower()
    if ext == ".csv":
        row_iter = dict_rows_from_csv(file_path)
    elif ext == ".xlsx":
        row_iter = dict_rows_from_xlsx(file_path)
    else:
        raise RuntimeError(f"Unsupported file type: {file_path.name}")

    rows_total = rows_loaded = rows_skipped = 0
    batch = 0

    try:
        for r in row_iter:
            rows_total += 1

            # Extract fields (case-insensitive pick)
            company = keypick(r, NAME_KEYS)
            phone_raw = keypick(r, PHONE_KEYS)
            email = keypick(r, EMAIL_KEYS)
            website = keypick(r, WEBSITE_KEYS)
            address = keypick(r, ADDR_KEYS)
            city = keypick(r, CITY_KEYS) or file_city
            state = keypick(r, STATE_KEYS) or file_state
            postal = keypick(r, ZIP_KEYS)
            contact = None  # extend later if you have a distinct contact column

            phone_e164 = norm_phone(phone_raw)
            tzname = tz_for(cur, city, state)

            row_data = {
                "company": (str(company).strip() if company not in (None, "") else None),
                "contact": contact,
                "email": (str(email).strip() if email not in (None, "") else None),
                "phone_raw": (str(phone_raw).strip() if phone_raw not in (None, "") else None),
                "phone_e164": phone_e164,
                "website": (str(website).strip() if website not in (None, "") else None),
                "address": (str(address).strip() if address not in (None, "") else None),
                "city": (str(city).strip() if city not in (None, "") else None),
                "state": (str(state).strip() if state not in (None, "") else None),
                "postal_code": (str(postal).strip() if postal not in (None, "") else None),
                "tz_name": tzname,
                "source_file": file_path.name,
                "source_row": json.dumps(r, ensure_ascii=False, default=str),
            }

            lead_id = upsert_lead_allow_nulls(cur, row_data)
            cur.execute("""
                INSERT INTO campaign_leads (campaign_id, lead_id, source_city, source_state, source_file)
                VALUES (%s,%s,%s,%s,%s)
                ON CONFLICT DO NOTHING
            """, (campaign_id, lead_id, row_data["city"], row_data["state"], file_path.name))

            rows_loaded += 1
            batch += 1
            if batch >= BATCH_SIZE:
                cur.connection.commit()
                print(f"    ... committed {rows_loaded}/{rows_total} from {file_path.name}")
                batch = 0

        # final commit for remainder
        cur.connection.commit()
        cur.execute("""
            UPDATE imports SET rows_total=%s, rows_loaded=%s, rows_skipped=%s,
                status='done', finished_at=now()
            WHERE id=%s
        """, (rows_total, rows_loaded, rows_skipped, imp_id))
        cur.connection.commit()
        print(f"  - DONE {file_path.name}: total={rows_total}, loaded={rows_loaded}, skipped={rows_skipped}")

    except Exception as e:
        cur.connection.rollback()
        cur.execute("UPDATE imports SET status='error', log=%s, finished_at=now() WHERE id=%s", (str(e), imp_id))
        cur.connection.commit()
        print(f"  ! ERROR {file_path.name}: {e}")

# ---------- Main ----------
def main():
    print(f"ROOT = {ROOT}")
    conn = psycopg2.connect(DB_URL)
    conn.autocommit = False
    with conn, conn.cursor(cursor_factory=psycopg2.extras.DictCursor) as cur:
        # Recursively find CSV + XLSX anywhere under ROOT
        files = []
        for pattern in ("**/*.csv", "**/*.xlsx"):
            files.extend(sorted(ROOT.rglob(pattern)))

        if not files:
            print("No CSV/XLSX files found under ROOT.")
            return

        print(f"Found {len(files)} files.")
        for fp in files:
            # Skip zero-byte files
            try:
                if fp.stat().st_size == 0:
                    print(f"  - SKIP zero-byte: {fp}")
                    continue
            except FileNotFoundError:
                continue

            print(f"\n== File => Campaign: {fp.name} => '{fp.stem}' ==")
            process_file(cur, fp)

    print("\nAll done.")

if __name__ == "__main__":
    main()
